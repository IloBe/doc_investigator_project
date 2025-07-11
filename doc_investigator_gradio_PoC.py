"""
A small gradio app for document investigation with error handling, typing and CLI logging.
This version includes a UI reset after evaluation and an analysis tab.
"""

# --- Imports ---
import os
import getpass
import sqlite3
import abc
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any, Type, Tuple

import gradio as gr
import fitz  # PyMuPDF for pdf handling
import docx
import openpyxl
import google.generativeai as genai


# --- Custom exceptions ---
class InvalidFileTypeException(Exception):
    """Custom exception raised for unsupported file types."""
    pass

# --- Config ---
@dataclass(frozen=True)
class Config:
    """Encapsulates all application configuration settings."""
    DB_FILE: str = "doc_investigator_prod.db"
    UNKNOWN_ANSWER: str = "Your request is unknown, associated information is not available. Please try again!"
    MAX_CONTEXT_CHARACTERS: int = 80000
    # Use a model name confirmed available via `genai.list_models()`
    # as checked with helper file 'check_google_models.py' of this directory
    LLM_MODEL_NAME: str = "gemini-2.5-pro"
    # Whenever you need to create a new Config object,
    # call this lambda function, which will return a brand new list. 
    # This correctly isolates the state of each object. Annotation:
    # A simple list is created only once, and will throw the following error:
    # ValueError: mutable default <class 'list'> for field SUPPORTED_FILE_TYPES is not allowed:
    # use default_factory
    TEMPERATURE: float = 0.2
    TOP_P: float = 0.95
    SUPPORTED_FILE_TYPES: List[str] = field(
        default_factory=lambda: ['.pdf', '.docx', '.txt', '.xlsx'])


# --- DB layer ---
class DatabaseManager:
    """Manages all interactions with the SQLite database."""
    
    def __init__(self, db_path: str):
        """
        Initializes the DatabaseManager.

        Args:
            db_path (str): The file path for the SQLite database.
        """
        self.db_path = db_path
        self._setup_database()

    def _setup_database(self) -> None:
        """Initializes the database and creates the interactions table if not present."""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS interactions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        timestamp TEXT NOT NULL,
                        prompt TEXT,
                        answer TEXT,
                        evaluation TEXT
                    )
                """)
                table_info = cursor.execute("PRAGMA table_info(interactions)").fetchall()
                column_names = [info[1] for info in table_info]
                # main LLM model properties for prompt output result
                if 'temperature' not in column_names:
                    cursor.execute("ALTER TABLE interactions ADD COLUMN temperature REAL")
                if 'top_p' not in column_names:
                    cursor.execute("ALTER TABLE interactions ADD COLUMN top_p REAL")
                
                conn.commit()
            print(f"Database '{self.db_path}' is ready with required attributes.")
        except sqlite3.Error as e:
            print(f"FATAL: Database setup failed: {e}")
            raise

    def log_interaction(self, prompt: str, answer: str, evaluation: str, temperature: float, top_p: float) -> None:
        """
        Logs a user interaction to the database.

        Args:
            prompt (str): The user's input prompt.
            answer (str): The LLM's generated answer.
            evaluation (str): The user's evaluation of the answer.
            temperature (float): The temperature value used for the LLM call.
            top_p (float): The top_p value used for the LLM call.
        
        Raises:
            sqlite3.Error: If there is an issue with the database transaction.
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO interactions (timestamp, prompt, answer, evaluation, temperature, top_p) VALUES (?, ?, ?, ?, ?, ?)",
                    (datetime.now().isoformat(), prompt, answer, evaluation, temperature, top_p)
                )
                print(f"User interaction with prompt '{prompt}' is logged to the database.")
                conn.commit()
        except sqlite3.Error as e:
            print(f"ERROR: Could not log interaction to database: {e}")
            raise  # Re-raise to be handled by the caller UI layer


# --- Doc layer (strategy design pattern) ---
class DocumentLoaderStrategy(abc.ABC):
    """Abstract base class for a document loading strategy."""
    @abc.abstractmethod
    def load(self, file_path: str) -> str:
        """Loads a file and returns its text content."""
        pass

class PDFLoaderStrategy(DocumentLoaderStrategy):
    """Strategy for loading text from PDF files."""
    def load(self, file_path: str) -> str:
        try:
            text = ""
            with fitz.open(file_path) as doc:
                for page in doc:
                    text += page.get_text()
            return text
        except (FileNotFoundError, fitz.fitz.PyMuPDFError) as e:
            print(f"Error loading PDF {file_path}: {e}")
            return f"[Error processing PDF: {os.path.basename(file_path)}]"

class DocxLoaderStrategy(DocumentLoaderStrategy):
    """Strategy for loading text from DOCX files."""
    def load(self, file_path: str) -> str:
        try:
            doc = docx.Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
        except (FileNotFoundError, Exception) as e: # python-docx can have various errors
            print(f"Error loading DOCX {file_path}: {e}")
            return f"[Error processing DOCX: {os.path.basename(file_path)}]"

class TextLoaderStrategy(DocumentLoaderStrategy):
    """Strategy for loading text from TXT files."""
    def load(self, file_path: str) -> str:
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except (FileNotFoundError, IOError) as e:
            print(f"Error loading TXT {file_path}: {e}")
            return f"[Error processing TXT: {os.path.basename(file_path)}]"

class ExcelLoaderStrategy(DocumentLoaderStrategy):
    """Strategy for loading text from Excel files."""
    def load(self, file_path: str) -> str:
        try:
            workbook = openpyxl.load_workbook(file_path)
            content = []
            for sheet_name in workbook.sheetnames:
                content.append(f"--- Sheet: {sheet_name} ---")
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    row_text = "\t".join([str(cell.value) if cell.value is not None else "" for cell in row])
                    content.append(row_text)
            return "\n".join(content)
        except (FileNotFoundError, Exception) as e:
            print(f"Error loading XLSX {file_path}: {e}")
            return f"[Error processing XLSX: {os.path.basename(file_path)}]"

class DocumentProcessor:
    """Context class that uses a strategy to process documents."""
    def __init__(self, supported_extensions: List[str]):
        """
        Initializes the DocumentProcessor.

        Args:
            supported_extensions (List[str]): A list of supported file extensions (e.g., ['.pdf', '.docx']).
        """
        self.supported_extensions = supported_extensions
        self._strategies: Dict[str, DocumentLoaderStrategy] = {
            '.pdf': PDFLoaderStrategy(),
            '.docx': DocxLoaderStrategy(),
            '.txt': TextLoaderStrategy(),
            '.xlsx': ExcelLoaderStrategy(),
            '.xls': ExcelLoaderStrategy()
        }
        print(f"DocumentProcessor initialized for types: {supported_extensions}")

    def validate_files(self, files: List[Any]) -> None:
            """
            Validates that all uploaded files have a supported extension.
    
            Args:
                files (List[Any]): A list of file objects from Gradio.
    
            Raises:
                InvalidFileTypeException: If any file in the list has an unsupported extension.
            """
            print(f"Validating {len(files)} files...")
            if not files:
                return  print("... Nothing to validate")
    
            for file in files:
                file_name = os.path.basename(file.name)
                _, file_extension = os.path.splitext(file_name)
                if file_extension.lower() not in self.supported_extensions:
                    error_msg = f"Unsupported file type: '{file_extension}' in file '{file_name}'."
                    print(error_msg) # Log the specific error before raising (warning)
                    raise InvalidFileTypeException(error_msg)
            
            print(f"All {len(files)} files passed validation.") # success

    def process_files(self, files: List[Any]) -> str:
        """
        Extracts text from a list of pre-validated Gradio file objects.
        Validation is passed for this list.

        Args:
            files (List[Any]): A list of validated file objects from Gradio.

        Returns:
            str: The combined text content of all files.
        """
        print(f"Processing {len(files)} validated files to extract text...")
        all_texts = []
        for file in files:
            file_path = file.name
            _, file_extension = os.path.splitext(file_path)
            
            # lookup is safe, assume validation has passed
            strategy = self._strategies[file_extension.lower()]
            text = strategy.load(file_path)
            all_texts.append(f"--- CONTENT FROM {os.path.basename(file_path)} ---\n{text}")
            
        print("Text extraction from all files complete.")  # success
        return "\n\n".join(all_texts)


# --- AI Gemini LLM layer ---
class GeminiService:
    """Handles all communication with the Google Gemini API."""
    def __init__(self, api_key: str, config: Config):
        self.config = config
        try:
            genai.configure(api_key=api_key)
            self.generation_config = genai.types.GenerationConfig(
                temperature=self.config.TEMPERATURE,
                top_p=self.config.TOP_P
            )
            self.model = genai.GenerativeModel(config.LLM_MODEL_NAME)
        except Exception as e:
            print(f"FATAL: Failed to configure Gemini API: {e}")
            raise

    def get_answer(self, full_text_context: str, user_prompt: str) -> str:
        """
        Generates an answer based on context and a prompt.

        Args:
            full_text_context (str): The context extracted from documents.
            user_prompt (str): The user's question.

        Returns:
            str: The generated answer from the LLM or an error message.
        """
        if not user_prompt:
            return "Please enter a prompt to continue."

        prompt_template = f"""
        You are a meticulous assistant. Your task is to answer the user's question based *ONLY* on the provided context.
        Do not use any external knowledge or make up information.
        If the information to answer the question is not in the context, you MUST respond with the exact phrase: '{self.config.UNKNOWN_ANSWER}'

        CONTEXT:
        ---
        {full_text_context}
        ---

        USER'S QUESTION:
        {user_prompt}

        ANSWER:
        """
        try:
            response = self.model.generate_content(prompt_template)
            if not response.parts:
                return "The model did not generate a response, possibly due to safety settings."
            return response.text.strip()
        except Exception as e:
            print(f"ERROR: Gemini API call failed: {e}")
            return "An error occurred while communicating with the AI model. Please check the logs."


# --- Gradio App ---
class AppUI:
    """Encapsulates the Gradio UI and its event handling logic."""
    def __init__(self, config: Config, db_manager: DatabaseManager,
                 doc_processor: DocumentProcessor, ai_service: GeminiService):
        self.config = config
        self.db_manager = db_manager
        self.doc_processor = doc_processor
        self.ai_service = ai_service
        self.app = self._build_ui()

    def _build_ui(self) -> gr.Blocks:
        """Constructs the entire Gradio user interface."""
        theme = gr.themes.Soft(
            primary_hue=gr.themes.colors.blue,
            secondary_hue=gr.themes.colors.sky,
            neutral_hue=gr.themes.colors.slate,
        ).set(
            body_background_fill="#F7F7F7",
            block_background_fill="white",
            block_border_width="1px",
            block_shadow="*shadow_drop_lg",
            block_radius="12px",
            input_background_fill="#F7F7F7",
        )

        with gr.Blocks(theme=theme, title="Document Investigation") as app:
            # State management
            state_full_text = gr.State(None)
            state_last_prompt = gr.State(None)
            state_last_answer = gr.State(None)
            
            # UI components and layout
            # header and instructions
            # added tabs for investigation and analysis
            gr.Markdown(
                """
                <div style="text-align: center;">
                    <h1><img src="https://em-content.zobj.net/source/google/387/magnifying-glass-tilted-left_1f50d.png" width="35" height="35" style="display:inline-block; vertical-align: middle;"> Document Investigation</h1>
                </div>
                """
            )
            
            with gr.Tabs():
                with gr.TabItem("Investigate"):
                    gr.Markdown(
                        """
                        ### How to use this tool:
                        1.  **Upload** one or more documents.
                        2.  **Ask a specific question** about the content of your documents.
                        3.  Click **"Investigate"** to get an answer from the AI.
                        4.  **Evaluate** the answer to help improve the system.
            
                        **Supported Document Types:**
                        - 📄 PDF (.pdf)
                        - 📝 Word (.docx)
                        - 📊 Excel (.xlsx)
                        - 🗒️ Text (.txt)
                        """
                    )

                    with gr.Row():
                        with gr.Column(scale=1):
                            # loose Gradio validation to allow custom handler to take care
                            file_uploader = gr.File(
                                label = "Step 1: Upload Your Documents",
                                file_count = "multiple",
                                file_types = None # allow any file type, custom handler manages validation
                            )
                            prompt_input = gr.Textbox(label="Step 2: Enter Your Prompt", lines=3)
                            submit_btn = gr.Button("Investigate", variant="primary")
                        with gr.Column(scale=2):
                            answer_output = gr.Markdown(
                                label="LLM Answer", value="Your answer will appear here...")
                            with gr.Column(visible=False) as evaluation_panel:
                                evaluation_radio = gr.Radio(
                                    ["✔️ Yes, the answer is helpful and accurate.", "❌ No, the answer is not helpful or inaccurate."],
                                    label="Step 3: Evaluate the Answer"
                                )
                                evaluation_button = gr.Button("Submit Evaluation")

                with gr.TabItem("Analyze Evaluations"):
                    gr.Markdown(
                        """
                        ### Analyzing Evaluation Data with Datasette

                        All your interactions and evaluations are stored in a local SQLite database (`doc_investigator_prod.db`).
                        Datasette provides an instant web interface to explore this data.

                        **Instructions:**
                        1.  Open a new terminal in the same project directory.
                        2.  Click the button below to copy the command, then paste it into your terminal and press Enter.
                        3.  A new browser tab will open with the Datasette interface where you can explore the `interactions` table.
                        """
                    )
                    datasette_command = f"datasette {self.config.DB_FILE} --open"
                    with gr.Row():
                        gr.Textbox(
                            value=datasette_command,
                            label="Command to run Datasette",
                            interactive=False,
                            show_copy_button=True
                        )

            # Event Handling
            # first for validation method
            file_uploader.upload(
                fn=self._handle_file_upload,
                inputs=[file_uploader],
                outputs=[file_uploader]
            )
            # second for button clicks
            submit_btn.click(
                fn=self._handle_investigation,
                inputs=[file_uploader,
                        prompt_input],
                outputs=[answer_output,
                         state_full_text,
                         state_last_prompt,
                         state_last_answer,
                         evaluation_panel]
            )

            evaluation_button.click(
                fn=self._handle_evaluation,
                inputs=[state_last_prompt,
                        state_last_answer,
                        evaluation_radio],
                # lists all components and states that need to be reset
                outputs=[file_uploader,
                         prompt_input,
                         answer_output,
                         evaluation_panel,
                         evaluation_radio,
                         state_full_text,
                         state_last_prompt,
                         state_last_answer]
            )

        return app

    def _handle_file_upload(self, files: List[Any]) -> Optional[List[Any]]:
        """
        Validates files immediately upon upload by calling the dedicated validation method.
        """
        if not files:
            return None # No files to validate, clear the component.

        try:
            self.doc_processor.validate_files(files)
            return files # Success: return the file list to keep it in the UI
        except InvalidFileTypeException as e:
            supported_types = ", ".join(self.config.SUPPORTED_FILE_TYPES)
            gr.Warning(f"{e} Supported types are: {supported_types}")
            return None # Failure: return None to clear the component and unlock UI

    def _handle_investigation(self, files: List[Any], prompt: str) -> tuple:
        """
        Orchestrates document processing and answer generation.
        The return tuple signature is expanded to allow resetting the file uploader on error.
        """
        try:
            if not files:
                raise gr.Error("Please upload at least one document to investigate.")

            full_text = self.doc_processor.process_files(files)   
            current_context = full_text
            if len(full_text) > self.config.MAX_CONTEXT_CHARACTERS:
                current_context = full_text[:self.config.MAX_CONTEXT_CHARACTERS]
            
            answer = self.ai_service.get_answer(current_context, prompt)
            show_eval = answer != self.config.UNKNOWN_ANSWER and "error" not in answer.lower()
            print("Successful investigation step.") # success
            return answer, current_context, prompt, answer, gr.update(visible=show_eval)

        except Exception as e:
            print("An unexpected error occurred during the investigation step.", exc_info=e) # exception
            gr.Error("An unexpected error occurred. Please check the console and try again.")
            return gr.update(), None, None, None, gr.update(visible=False)


    def _handle_evaluation(self, prompt: str, answer: str, choice: str) -> Tuple:
            """
            Handles submission of the evaluation, logs it, and resets the UI to its initial state.
    
            Returns:
                A tuple of Gradio updates and None values to reset all relevant components and state.
                The length and order of this tuple MUST match the 'outputs' list in the click event.
            """
            if not choice:
                gr.Warning("Please select an evaluation option before submitting.")
                # Return a tuple of no-op updates to preserve the UI state.
                return (gr.update(), gr.update(), gr.update(),
                        gr.update(visible=True), gr.update(), gr.update(), gr.update())
            
            try:
                self.db_manager.log_interaction(
                    prompt, answer, choice,
                    self.config.TEMPERATURE,
                    self.config.TOP_P
                )
                print((f"Evaluation ('{choice[:5]}...') submitted and logged successfully."))
                gr.Info("Evaluation saved! The interface has been reset for next run.")
                
                # full UI reset, order must match the 'outputs' list
                return (
                    # Reset of ...
                    gr.update(value=None),      # file_uploader
                    gr.update(value=""),        # prompt_input
                    gr.update(value="Your answer will appear here..."), # answer_output
                    gr.update(visible=False),   # evaluation_panel
                    gr.update(value=None),      # evaluation_radio
                    None,                       # state_full_text
                    None,                       # state_last_prompt
                    None                        # state_last_answer
                )
            except sqlite3.Error:
                print(f"gradio handle evaluation: Failed to save evaluation due to a database error")
                gr.Error("Failed to save evaluation due to a database error.")
                # Keep the panel open and state intact so the user can try again.
                return (gr.update(), gr.update(), gr.update(),
                        gr.update(visible=True), gr.update(), gr.update(), gr.update(), gr.update())

            except Exception as e:
                print("Failed to handle evaluation submission.", exc_info=e)
                gr.Error("Failed to save evaluation due to a system error. Please try again.")
                # Keep the panel open and state intact so the user can try again.
                return (gr.update(), gr.update(), gr.update(),
                        gr.update(visible=True), gr.update(), gr.update(), gr.update(), gr.update())

    def launch(self):
        """Launches the Gradio application."""
        self.app.launch(debug=True)

# --- Execution ---       
def main():
    """Main function to initialize and run the application."""
    config = Config()

    api_key = os.environ.get('GOOGLE_API_KEY')
    if not api_key:
        try:
            api_key = getpass.getpass('Enter your Google Gemini API Key: ')
        except Exception as e:
            print(f"Could not read API key: {e}")
            return

    try:
        db_manager = DatabaseManager(config.DB_FILE)
        doc_processor = DocumentProcessor(config.SUPPORTED_FILE_TYPES)
        ai_service = GeminiService(api_key, config)
    except Exception as e:
        print(f"Application failed to initialize. Aborting. Error: {e}")
        return

    app_ui = AppUI(config, db_manager, doc_processor, ai_service)
    app_ui.launch()

if __name__ == "__main__":
    main()